import openpyxl
from openpyxl.styles import PatternFill
import os
from utils import compare_models, MODELS

def warn(*args, **kwargs):
    pass
import warnings
warnings.warn = warn

script_dir = os.path.dirname(os.path.abspath(__file__)) #getting location in sub directory
save_dir = os.path.join(script_dir, "Tournament Bracket.xlsx")

wb = openpyxl.load_workbook(os.path.join(script_dir, "Tournament Template.xlsx"))
sheet = wb.active
round_to_column = {1: 3, 2: 6, 3: 9, 4: 13, 5: 17}
TOP_LEFT_ROW = 9

def get_top_row(round_number):
    return sum((2 ** (i-1)) for i in range(1, round_number+1)) + TOP_LEFT_ROW #getting the first square

def update_round(round_number, winners):
    next_column = round_to_column[round_number + 1]
    start_row = get_top_row(round_number)
    for i, winner in enumerate(winners):
        row = start_row + (i * (2 ** (round_number+1)))
        cell = sheet.cell(row=row, column=next_column)
        cell.value = winner
    wb.save(save_dir)

def play_round(round_number):
    matches = 16 / (2 ** round_number)
    row, column = get_top_row(round_number - 1), round_to_column[round_number]
    winners = []
    while matches > 0:
        model1 = sheet.cell(row=row, column=column).value
        row += (2 ** (round_number))
        model2 = sheet.cell(row=row, column=column).value
        row += (2 ** (round_number))
        winners.append(compare_models(model1, model2, round_number))
        matches -= 1
    update_round(round_number, winners)

for i in range(1, 5):
    play_round(i)
    if input("Continue? (y/n)") == "n":
        break
    print("=====================================")



